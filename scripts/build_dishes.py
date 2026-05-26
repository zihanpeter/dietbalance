"""从 OneDrive 中的菜品 Excel 表生成应用使用的 ``data/dishes.json``。

Excel 结构（首行为大标题，第二行为列头）::

    菜品名称 | 常规食量(100g)热量(千卡) | 碳水化合物占比 | 蛋白质占比 | 脂肪占比 | 菜品特点

百分比单元格存的是诸如 ``"70.0%"`` 的字符串，本脚本会把它们解析成 ``70.0``。
脚本只在数据源更新时手动运行一次。
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

XLSX_PATH = Path(r"C:\Users\luzih\OneDrive\食堂菜品营养热量与特点表.xlsx")
OUT_JSON = Path(__file__).resolve().parent.parent / "data" / "dishes.json"

sys.stdout.reconfigure(encoding="utf-8")


def _to_pct(value: object) -> float:
    """把 ``"70.0%"`` 之类的单元格解析为 ``70.0``。"""
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        # openpyxl 偶尔会把百分比读成 0.7
        return float(value) * 100 if value <= 1 else float(value)
    text = str(value).strip().replace("％", "%")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else 0.0


def _to_kcal(value: object) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return int(round(float(value)))
    match = re.search(r"-?\d+(?:\.\d+)?", str(value))
    return int(round(float(match.group(0)))) if match else 0


def main() -> None:
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb.active
    print(f"Loading sheet: {ws.title} ({ws.max_row} rows)")

    dishes: list[dict[str, object]] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        if not name or name.startswith("菜品名称"):
            continue

        dishes.append(
            {
                "name": name,
                "kcal": _to_kcal(row[1]),
                "carb_pct": round(_to_pct(row[2]), 2),
                "protein_pct": round(_to_pct(row[3]), 2),
                "fat_pct": round(_to_pct(row[4]), 2),
                "features": (str(row[5]).strip() if row[5] else ""),
            }
        )

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(
        json.dumps(dishes, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"Wrote {len(dishes)} dishes -> {OUT_JSON}")


if __name__ == "__main__":
    main()
